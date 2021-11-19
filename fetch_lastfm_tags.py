import pylast
import openpyxl
import os
import os.path
import operator
import json

API_KEY = ""  
API_SECRET = ""
USERNAME = ""
PASSWORD_HASH = pylast.md5("")



def setupLast(API_KEY, API_SECRET, username, password_hash):
    network = pylast.LastFMNetwork(
    api_key=API_KEY,
    api_secret=API_SECRET,
    username=username,
    password_hash=password_hash,
)

    return network

network = setupLast(API_KEY, API_SECRET, USERNAME, PASSWORD_HASH)

def findGenreByTrackOrArtist(type, artist, track):
    allGenreList = []
    genreList = []
    totalWeight = 0
    runningWeight = 0
    if(type == 'TR'):
        topItems = track.get_top_tags(limit=5)
        if(len(topItems) == 0):
            raise ValueError("No info found for this track")
        else:
            for topItem in topItems:
                allGenreList.append([topItem.item.get_name(), topItem.weight])
                #print(topItem.item.get_name(), topItem.weight)
            for i in range(0,len(allGenreList)):
                totalWeight += int(allGenreList[i][1])
            for i in range(0, len(allGenreList)):
                if(runningWeight/totalWeight <= 0.95):
                    runningWeight += int(allGenreList[i][1])
                    genreList.append([allGenreList[i][0], allGenreList[i][1]])
                else:
                    break
            return genreList
    elif(type == 'AR'):
        topArtistItems = artist.get_top_tags(limit=5)
        if(len(topArtistItems) == 0):
            raise ValueError("No info found for this artist")
        else:
            for topItem in topArtistItems:
                    allGenreList.append([topItem.item.get_name(), topItem.weight])
                    #print(topItem.item.get_name(), topItem.weight)
            for i in range(0,len(allGenreList)):
                totalWeight += int(allGenreList[i][1])
            for i in range(0, len(allGenreList)):
                if(runningWeight/totalWeight <= 0.95):
                    runningWeight += int(allGenreList[i][1])
                    genreList.append([allGenreList[i][0], allGenreList[i][1]])
                else:
                    break
            return genreList

def updateTagDictionary(tagList, dictionary,normalized_dictionary):
    totalSum = 0
    for tag in tagList:
        if dictionary.get(tag[0]) is not None:
            dictionary[tag[0]] += int(tag[1])
        else:
            dictionary[tag[0]] = int(tag[1])
    for tag in dictionary.keys():
        totalSum += dictionary[tag]
    for tag in dictionary.keys():
        normalized_dictionary[tag] = dictionary[tag] / totalSum


def generateTopTags(norm_dict):
    GTZAN_TOP_TAGS = {}
    TOTAL_NORMALIZED_TAG_COUNT  = 0
    for tag in norm_dict.keys():
        TOTAL_NORMALIZED_TAG_COUNT += norm_dict[tag]
    print(f"TOTAL TAG COUNT {TOTAL_NORMALIZED_TAG_COUNT}")
    running_norm_tag_count = 0
    for tag in norm_dict.keys():
        if(running_norm_tag_count/TOTAL_NORMALIZED_TAG_COUNT <= 0.95):
            running_norm_tag_count += norm_dict[tag]
            GTZAN_TOP_TAGS[tag] = norm_dict[tag]
        else:
            break
    return GTZAN_TOP_TAGS

def normalizeExcerptTags(tagList):
    TAGS = {}
    totalCount = 0
    for tag in tagList:
        totalCount += int(tag[1])
    for tag in tagList:
        TAGS[tag[0]] = int(tag[1]) / totalCount
    return TAGS
    


def findGenre(PATH):
    STARTING_ROW = 3
    ARTIST_COLUMN = 3
    TOTAL_ROWS_COLUMN = 5
    LASTFM_GENRE_COLUMN = 6
    LASTFM_INFO_COLUMN = 7
    NORMALIZED_TAG_COUNT_COLUMN = 8
    GTZAN_TOP_TAGS_COLUMN = 9
    TRACK_COLUMN = 2
    GTZAN_ALL_TOP_TAGS = {}
    GTZAN_ALL_EXCERPT_TAGS = {}
    GENRE_SCORES = {}
    for root, dirs, files in os.walk(PATH):
        for File in files:
            GTZAN_EXCERPT_TAGS = []
            try:
                TAG_DICTIONARY = dict()
                NORMALIZED_TAG_DICTIONARY = dict()
                root = root.replace("\\","/")
                dataFilePath = f"{root}{File}"
                print(f"Now Querying Last.fm with all the entries in -- {dataFilePath}")
                wb = openpyxl.load_workbook(dataFilePath)
                sheet = wb.active
                totalQueried = 0
                totalQueriedByTrack = 0
                totalQueriedByArtist = 0
                lastfm_info_cell = sheet.cell(row=STARTING_ROW, column=LASTFM_INFO_COLUMN)
                totalRows = int(sheet.cell(row=STARTING_ROW, column=TOTAL_ROWS_COLUMN).value)+1
                for row in range(STARTING_ROW, totalRows):
                    try:
                        lastfm_genre_cell = sheet.cell(row=row, column=LASTFM_GENRE_COLUMN)
                        artist = sheet.cell(row=row, column=ARTIST_COLUMN).value
                        track = sheet.cell(row=row, column=TRACK_COLUMN).value
                        print(artist, track)
                        lastfm_artist = network.get_artist(artist)
                        lastfm_track = network.get_track(title=track, artist=artist)
                        lastfm_genreByTrack = findGenreByTrackOrArtist('TR', lastfm_artist, lastfm_track)
                        normalizedExcerptTags = normalizeExcerptTags(lastfm_genreByTrack)
                        GTZAN_EXCERPT_TAGS.append(normalizedExcerptTags)
                        lastfm_genre_cell.value = f'TRACK - {normalizedExcerptTags}'
                        #wb.save(dataFilePath)
                        totalQueried += 1
                        totalQueriedByTrack += 1
                        print(f"FOUND GENRE (TRACK) [{lastfm_artist} - {lastfm_track}] -- {lastfm_genreByTrack}")
                        updateTagDictionary(lastfm_genreByTrack, TAG_DICTIONARY, NORMALIZED_TAG_DICTIONARY)
                        if(row == totalRows - 1):
                            NORMALIZED_TAG_DICTIONARY = dict(sorted(NORMALIZED_TAG_DICTIONARY.items(), key=operator.itemgetter(1),reverse=True))
                            GTZAN_Top_Tags = generateTopTags(NORMALIZED_TAG_DICTIONARY)
                            GTZAN_ALL_TOP_TAGS[File] = GTZAN_Top_Tags
                            GTZAN_ALL_EXCERPT_TAGS[File] = GTZAN_EXCERPT_TAGS
                            sheet.cell(row=STARTING_ROW, column=GTZAN_TOP_TAGS_COLUMN).value = f'{GTZAN_Top_Tags}'
                        sheet.cell(row=STARTING_ROW, column=NORMALIZED_TAG_COUNT_COLUMN).value = f'{NORMALIZED_TAG_DICTIONARY}'
                        wb.save(dataFilePath)
                    except:
                        try:
                            lastfm_genreByArtist = findGenreByTrackOrArtist('AR', lastfm_artist, lastfm_track)
                            normalizedExcerptTags = normalizeExcerptTags(lastfm_genreByArtist)
                            GTZAN_EXCERPT_TAGS.append(normalizedExcerptTags)
                            lastfm_genre_cell.value = f'ARTIST - {normalizedExcerptTags}'
                            #wb.save(dataFilePath)
                            totalQueried += 1
                            totalQueriedByArtist += 1
                            print(f"FOUND GENRE (ARTIST) [{lastfm_artist} - {lastfm_track}] -- {lastfm_genreByArtist}")
                            updateTagDictionary(lastfm_genreByArtist, TAG_DICTIONARY, NORMALIZED_TAG_DICTIONARY)
                            if(row == totalRows - 1):
                                NORMALIZED_TAG_DICTIONARY = dict(sorted(NORMALIZED_TAG_DICTIONARY.items(), key=operator.itemgetter(1),reverse=True))
                                GTZAN_Top_Tags = generateTopTags(NORMALIZED_TAG_DICTIONARY)
                                GTZAN_ALL_TOP_TAGS[File] = GTZAN_Top_Tags
                                GTZAN_ALL_EXCERPT_TAGS[File] = GTZAN_EXCERPT_TAGS
                                sheet.cell(row=STARTING_ROW, column=GTZAN_TOP_TAGS_COLUMN).value = f'{GTZAN_Top_Tags}'
                            sheet.cell(row=STARTING_ROW, column=NORMALIZED_TAG_COUNT_COLUMN).value = f'{NORMALIZED_TAG_DICTIONARY}'
                            wb.save(dataFilePath)
                        except:
                            lastfm_genre_cell.value = 'N/A'
                            GTZAN_EXCERPT_TAGS.append({})   
                            if(row == totalRows - 1):
                                NORMALIZED_TAG_DICTIONARY = dict(sorted(NORMALIZED_TAG_DICTIONARY.items(), key=operator.itemgetter(1),reverse=True))
                                GTZAN_Top_Tags = generateTopTags(NORMALIZED_TAG_DICTIONARY)
                                GTZAN_ALL_TOP_TAGS[File] = GTZAN_Top_Tags
                                GTZAN_ALL_EXCERPT_TAGS[File] = GTZAN_EXCERPT_TAGS
                                sheet.cell(row=STARTING_ROW, column=GTZAN_TOP_TAGS_COLUMN).value = f'{GTZAN_Top_Tags}'
                            wb.save(dataFilePath)
                            print("No information found")
                print(GTZAN_ALL_TOP_TAGS)
                print(GTZAN_ALL_EXCERPT_TAGS)
                print(f"Total Song Genres Found - {totalQueried} / {sheet.max_row - STARTING_ROW + 1}")
                print(f"Total Song Genres Found (By Track) - {totalQueriedByTrack}")
                print(f"Total Song Genres Found (By Artist) - {totalQueriedByArtist}")
                lastfm_info_cell.value = f'Total Song Genres Found - {totalQueried} / {sheet.max_row - STARTING_ROW + 1}, Found By Track - {totalQueriedByTrack}, Found By Artist - {totalQueriedByArtist}'
                wb.save(dataFilePath)
            except:
                wb.save(dataFilePath)
                print("Error Opening File")
    
    #SAVING GENRE SCORES FOR EACH EXCERPT
    for excerptTag, excerptTagSetList in GTZAN_ALL_EXCERPT_TAGS.items():
        print(f'{excerptTag}')
        count = 0
        SCORE_TEMP = dict()
        for tagSet in excerptTagSetList:
            SCORE_TAG_TEMP = dict()
            if tagSet:
                for labelTag, labelTagDict in GTZAN_ALL_TOP_TAGS.items():
                    genre_score = 0
                    for tag, tagScore in tagSet.items():
                        if(tag in labelTagDict):
                            genre_score += tagScore * labelTagDict[tag]
                    SCORE_TAG_TEMP[labelTag] = genre_score
                SCORE_TEMP[count] = SCORE_TAG_TEMP
                count += 1
            else:
                SCORE_TEMP[count] = 'N/A'
                count += 1
        with open(f'../fingerprinter/GTZAN-Last.fm Genres/{excerptTag}.txt', 'w') as json_file:
            json.dump(SCORE_TEMP, json_file)
        print(SCORE_TEMP)
            

            

    
    
            

                        
                
                
        
        

            
    
findGenre("../fingerprinter/GTZAN-Last.fm Genres/")
