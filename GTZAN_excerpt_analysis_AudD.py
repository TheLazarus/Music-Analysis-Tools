import requests
import os
import os.path
import json
import openpyxl

path = ''

wb = openpyxl.load_workbook(path)
sheet = wb.active 

data = {
    'api_token': '',
}
PATH = ''
def getAudioInformation():
    sno = 0
    ROW = 2
    tags = []
    for root, dirs, files in os.walk(PATH):
        for File in files:
            try:
                root = root.replace("\\","/")
                files = {
                        'file': open(f'{root}/{File}', 'rb'),
                }
                result = requests.post('https://api.audd.io/', data=data, files=files)
                result_json = json.loads(result.text)
                metadata = {
                'sno': sno,
                'artist': result_json['result']['artist'],
                'title': result_json['result']['title'],
                'album': result_json['result']['album']
                }
                print(metadata)
                songid = sheet.cell(row=ROW, column=1)
                track = sheet.cell(row=ROW, column=2)
                artist = sheet.cell(row=ROW, column=3)
                album = sheet.cell(row=ROW, column=4)
                songid.value = sno
                track.value = result_json['result']['title']
                artist.value = result_json['result']['artist']
                album.value = result_json['result']['album']
                tags.append(metadata)
                sno += 1 
                ROW += 1
                wb.save(path) 
            except:
                sno += 1
                continue
    return tags

print(getAudioInformation())
